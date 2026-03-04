"""Glossary management endpoints."""

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse, Response
from typing import List, Dict
from sqlalchemy.orm import Session
import uuid
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.term import (
    Term,
    TermUpdate,
    TermStatus,
    GlossaryResponse,
    GlossaryStats,
    SourceReport,
    SourceReportItem,
    GlossarySessionCreate,
    GlossarySessionResponse,
    ManualTermCreate,
)
from db.database import get_db
from db import models

router = APIRouter(prefix="/glossary", tags=["glossary"])


@router.get("/{document_id}", response_model=GlossaryResponse)
async def get_glossary(
    document_id: str,
    status: str = Query("all", description="Filter by status: all, pending, approved, edited, rejected"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """
    Get glossary terms for a document.

    Args:
        document_id: Document ID
        status: Filter by term status
        page: Page number
        per_page: Items per page
        db: Database session

    Returns:
        GlossaryResponse with terms and statistics
    """
    # Get all terms for this document from database
    query = db.query(models.Term).filter(models.Term.document_id == document_id)

    all_terms = query.all()

    # Calculate stats
    stats = GlossaryStats(
        total=len(all_terms),
        pending=len([t for t in all_terms if t.status == "pending"]),
        approved=len([t for t in all_terms if t.status == "approved"]),
        edited=len([t for t in all_terms if t.status == "edited"]),
        rejected=len([t for t in all_terms if t.status == "rejected"]),
        # Source breakdown
        from_hudoc=len([t for t in all_terms if t.source_type == "hudoc"]),
        from_curia=len([t for t in all_terms if t.source_type == "curia"]),
        from_iate=len([t for t in all_terms if t.source_type == "iate"]),
        from_tm_exact=len([t for t in all_terms if t.source_type == "tm_exact"]),
        from_tm_fuzzy=len([t for t in all_terms if t.source_type == "tm_fuzzy"]),
        from_proposed=len([t for t in all_terms if t.source_type == "proposed"]),
    )

    # FIXED: Create a NEW query for paginated results (don't reuse executed query)
    paginated_query = db.query(models.Term).filter(models.Term.document_id == document_id)

    # Filter by status if requested
    if status != "all":
        paginated_query = paginated_query.filter(models.Term.status == status)

    # Pagination
    doc_terms = paginated_query.offset((page - 1) * per_page).limit(per_page).all()

    # Convert DB models to Pydantic models
    terms_list = []
    for t in doc_terms:
        # Extract sources from references
        sources_list = []
        context = None
        if t.references and isinstance(t.references, dict):
            context = t.references.get("context")

            # Check for case_law_references (list of source objects from CaseLawResearcher)
            case_law_refs = t.references.get("case_law_references", [])
            if case_law_refs:
                for ref in case_law_refs:
                    if isinstance(ref, dict):
                        source_info = {
                            "source_type": ref.get("source", "unknown"),
                            "case_name": ref.get("case_name"),
                            "url": ref.get("url"),
                            "context": ref.get("context"),
                        }
                        sources_list.append(source_info)

        term_dict = {
            "id": t.id,
            "document_id": t.document_id,
            "source_term": t.source_term,
            "target_term": t.target_term,
            "original_proposal": t.original_proposal,
            "source_type": t.source_type,
            "confidence": t.confidence,
            "references": [],  # Database stores as dict, but Pydantic expects list - context extracted separately
            "status": t.status,
            "context": context,
            "sources": sources_list,
            "created_at": t.created_at,
            "updated_at": t.updated_at,
        }
        terms_list.append(Term(**term_dict))

    return GlossaryResponse(
        total=query.count() if status != "all" else len(all_terms),
        stats=stats,
        terms=terms_list,
    )


@router.post("/{document_id}/terms", response_model=Term)
async def create_manual_term(
    document_id: str,
    term_data: ManualTermCreate,
    db: Session = Depends(get_db),
):
    """
    Create a manual term (added by user).

    Args:
        document_id: Document ID
        term_data: Term data (source_term, target_term, context, notes)
        db: Database session

    Returns:
        Created term
    """
    # Verify document exists
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Check if term already exists
    existing_term = db.query(models.Term).filter(
        models.Term.document_id == document_id,
        models.Term.source_term == term_data.source_term
    ).first()

    if existing_term:
        raise HTTPException(
            status_code=400,
            detail=f"Termin '{term_data.source_term}' juz istnieje w glosariuszu"
        )

    # Create new term
    new_term = models.Term(
        id=str(uuid.uuid4()),
        document_id=document_id,
        source_term=term_data.source_term,
        target_term=term_data.target_term,
        original_proposal=term_data.target_term,
        source_type="manual",
        confidence=1.0,
        status="approved",  # Manual terms are auto-approved
        references={"context": term_data.context, "notes": term_data.notes} if term_data.context or term_data.notes else None,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )

    db.add(new_term)
    db.commit()
    db.refresh(new_term)

    return Term(
        id=new_term.id,
        document_id=new_term.document_id,
        source_term=new_term.source_term,
        target_term=new_term.target_term,
        original_proposal=new_term.original_proposal,
        source_type=new_term.source_type,
        confidence=new_term.confidence,
        status=new_term.status,
        context=term_data.context,
        references=[],
        sources=[],
        created_at=new_term.created_at,
        updated_at=new_term.updated_at,
    )


@router.put("/{document_id}/{term_id}", response_model=Term)
async def update_term(document_id: str, term_id: str, update: TermUpdate, db: Session = Depends(get_db)):
    """
    Update a term (approve, edit, or reject).

    Args:
        document_id: Document ID
        term_id: Term ID
        update: Term update data
        db: Database session

    Returns:
        Updated term
    """
    # Find term in database
    term = db.query(models.Term).filter(
        models.Term.id == term_id,
        models.Term.document_id == document_id
    ).first()

    if not term:
        raise HTTPException(status_code=404, detail="Term not found")

    # Save original proposal if this is the first edit
    if term.original_proposal is None and update.status == TermStatus.EDITED:
        term.original_proposal = term.target_term

    # Update term
    term.target_term = update.target_term
    term.status = update.status.value
    term.updated_at = datetime.now()

    db.commit()
    db.refresh(term)

    # Extract sources from references
    sources_list = []
    context = None
    if term.references and isinstance(term.references, dict):
        context = term.references.get("context")

        # Check for case_law_references
        case_law_refs = term.references.get("case_law_references", [])
        if case_law_refs:
            for ref in case_law_refs:
                if isinstance(ref, dict):
                    source_info = {
                        "source_type": ref.get("source", "unknown"),
                        "case_name": ref.get("case_name"),
                        "url": ref.get("url"),
                        "context": ref.get("context"),
                    }
                    sources_list.append(source_info)

    return Term(
        id=term.id,
        document_id=term.document_id,
        source_term=term.source_term,
        target_term=term.target_term,
        original_proposal=term.original_proposal,
        source_type=term.source_type,
        confidence=term.confidence,
        references=[],  # Database stores as dict, but Pydantic expects list - context extracted separately
        status=term.status,
        context=context,
        sources=sources_list,
        created_at=term.created_at,
        updated_at=term.updated_at,
    )


@router.post("/{document_id}/{term_id}/apply-to-translation")
async def apply_term_to_translation(
    document_id: str,
    term_id: str,
    db: Session = Depends(get_db)
):
    """
    Apply updated term translation to all segments in the document.

    This endpoint finds all occurrences of the term's old translation in segments
    and replaces them with the new translation.

    Args:
        document_id: Document ID
        term_id: Term ID
        db: Database session

    Returns:
        Dict with number of segments updated
    """
    # Find the term
    term = db.query(models.Term).filter(
        models.Term.id == term_id,
        models.Term.document_id == document_id
    ).first()

    if not term:
        raise HTTPException(status_code=404, detail="Term not found")

    # Get old and new translations
    new_translation = term.target_term
    old_translation = term.original_proposal if term.original_proposal else term.target_term

    if old_translation == new_translation:
        return {"segments_updated": 0, "message": "No change in translation"}

    # Get all translated segments for this document
    segments = db.query(models.Segment).filter(
        models.Segment.document_id == document_id,
        models.Segment.target_text.isnot(None),
        models.Segment.target_text != ""
    ).all()

    updated_count = 0

    # Replace old translation with new in each segment
    for segment in segments:
        if old_translation in segment.target_text:
            # Simple case-sensitive replacement
            segment.target_text = segment.target_text.replace(old_translation, new_translation)
            segment.updated_at = datetime.now()
            updated_count += 1

    db.commit()

    return {
        "segments_updated": updated_count,
        "old_translation": old_translation,
        "new_translation": new_translation,
        "message": f"Updated {updated_count} segment(s)"
    }


@router.post("/{document_id}/approve-all")
async def approve_all_pending(document_id: str, db: Session = Depends(get_db)):
    """
    Approve all pending terms for a document.

    Args:
        document_id: Document ID
        db: Database session

    Returns:
        Count of approved terms
    """
    # Get all pending terms for this document
    pending_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id,
        models.Term.status == "pending"
    ).all()

    count = 0
    for term in pending_terms:
        term.status = "approved"
        term.updated_at = datetime.now()
        count += 1

    db.commit()

    return {"message": f"Approved {count} terms", "count": count}


@router.get("/{document_id}/sources-report", response_model=SourceReport)
async def get_sources_report(document_id: str, db: Session = Depends(get_db)):
    """
    Get detailed report of term sources with case information.

    Args:
        document_id: Document ID
        db: Database session

    Returns:
        SourceReport with terms grouped by source type
    """
    # Get all terms for this document
    all_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id
    ).all()

    # Helper function to extract case info from references
    def extract_case_info(references):
        if not references:
            return None, None, None

        if isinstance(references, dict):
            case_name = references.get("case_name")
            case_url = references.get("url")
            context = references.get("context")
            return case_name, case_url, context
        return None, None, None

    # Group terms by source type
    hudoc_terms = []
    curia_terms = []
    iate_terms = []
    tm_exact_terms = []
    tm_fuzzy_terms = []
    proposed_terms = []

    for t in all_terms:
        case_name, case_url, context = extract_case_info(t.references)

        item = SourceReportItem(
            source_term=t.source_term,
            target_term=t.target_term,
            source_type=t.source_type,
            case_name=case_name,
            case_url=case_url,
            context=context,
            status=t.status,
        )

        if t.source_type == "hudoc":
            hudoc_terms.append(item)
        elif t.source_type == "curia":
            curia_terms.append(item)
        elif t.source_type == "iate":
            iate_terms.append(item)
        elif t.source_type == "tm_exact":
            tm_exact_terms.append(item)
        elif t.source_type == "tm_fuzzy":
            tm_fuzzy_terms.append(item)
        elif t.source_type == "proposed":
            proposed_terms.append(item)

    return SourceReport(
        hudoc_terms=hudoc_terms,
        curia_terms=curia_terms,
        iate_terms=iate_terms,
        tm_exact_terms=tm_exact_terms,
        tm_fuzzy_terms=tm_fuzzy_terms,
        proposed_terms=proposed_terms,
    )


@router.get("/{document_id}/export/all/xlsx")
async def export_all_terms_xlsx(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport wszystkich terminów do XLSX (przed zatwierdzeniem).

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Plik XLSX ze wszystkimi terminami
    """
    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Pobierz wszystkie terminy
    all_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id
    ).order_by(models.Term.source_term).all()

    if not all_terms:
        raise HTTPException(status_code=404, detail="No terms found for this document")

    # Utwórz workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary - All Terms"

    # Style dla nagłówka
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style dla danych
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    # Header
    headers = ['Source Term (EN)', 'Target Term (PL)', 'Source Type', 'Status', 'Confidence', 'Case Name', 'Context']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_num, term in enumerate(all_terms, 2):
        case_name = ""
        context = ""

        if term.references and isinstance(term.references, dict):
            case_name = term.references.get("case_name", "")
            context = term.references.get("context", "")

        # Status color coding
        status_fill = None
        if term.status == "approved":
            status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        elif term.status == "edited":
            status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        elif term.status == "rejected":
            status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

        data = [
            term.source_term,
            term.target_term or "",
            term.source_type,
            term.status,
            f"{term.confidence:.2f}",
            case_name,
            context
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = cell_alignment

            # Apply status color to Status column
            if col_num == 4 and status_fill:
                cell.fill = status_fill

    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # Source Term
    ws.column_dimensions['B'].width = 30  # Target Term
    ws.column_dimensions['C'].width = 15  # Source Type
    ws.column_dimensions['D'].width = 12  # Status
    ws.column_dimensions['E'].width = 12  # Confidence
    ws.column_dimensions['F'].width = 25  # Case Name
    ws.column_dimensions['G'].width = 50  # Context

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"glossary_all_{document_id}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/{document_id}/export/all/html")
async def export_all_terms_html(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport wszystkich terminów do HTML (przed zatwierdzeniem).

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Plik HTML ze wszystkimi terminami
    """
    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Pobierz wszystkie terminy
    all_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id
    ).order_by(models.Term.source_term).all()

    if not all_terms:
        raise HTTPException(status_code=404, detail="No terms found for this document")

    # Statystyki
    stats = {
        'total': len(all_terms),
        'approved': len([t for t in all_terms if t.status == 'approved']),
        'edited': len([t for t in all_terms if t.status == 'edited']),
        'pending': len([t for t in all_terms if t.status == 'pending']),
        'rejected': len([t for t in all_terms if t.status == 'rejected']),
    }

    # Generate HTML
    html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Glossary - All Terms - {db_document.filename}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        .meta {{
            color: #7f8c8d;
            margin-bottom: 20px;
            font-size: 14px;
        }}
        .stats {{
            display: flex;
            gap: 15px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }}
        .stat {{
            background: #ecf0f1;
            padding: 10px 20px;
            border-radius: 5px;
            font-size: 14px;
        }}
        .stat strong {{
            color: #2c3e50;
            font-size: 18px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #3498db;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 10px;
            border: 1px solid #ddd;
            vertical-align: top;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f0f8ff;
        }}
        .status-approved {{
            background-color: #d4edda !important;
            color: #155724;
            font-weight: 600;
        }}
        .status-edited {{
            background-color: #fff3cd !important;
            color: #856404;
            font-weight: 600;
        }}
        .status-pending {{
            background-color: #cce5ff !important;
            color: #004085;
        }}
        .status-rejected {{
            background-color: #f8d7da !important;
            color: #721c24;
        }}
        .source-type {{
            font-size: 11px;
            padding: 3px 8px;
            border-radius: 3px;
            display: inline-block;
            background: #e9ecef;
        }}
        .confidence {{
            font-weight: 600;
            color: #27ae60;
        }}
        .context {{
            font-size: 12px;
            color: #555;
            max-width: 400px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📚 Glossary - All Terms</h1>
        <div class="meta">
            <strong>Document:</strong> {db_document.filename}<br>
            <strong>Document ID:</strong> {document_id}<br>
            <strong>Exported:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>

        <div class="stats">
            <div class="stat">📊 <strong>{stats['total']}</strong> Total Terms</div>
            <div class="stat" style="background: #d4edda;">✓ <strong>{stats['approved']}</strong> Approved</div>
            <div class="stat" style="background: #fff3cd;">✏️ <strong>{stats['edited']}</strong> Edited</div>
            <div class="stat" style="background: #cce5ff;">⏳ <strong>{stats['pending']}</strong> Pending</div>
            <div class="stat" style="background: #f8d7da;">✗ <strong>{stats['rejected']}</strong> Rejected</div>
        </div>

        <table>
            <thead>
                <tr>
                    <th style="width: 20%;">Source Term (EN)</th>
                    <th style="width: 20%;">Target Term (PL)</th>
                    <th style="width: 10%;">Source</th>
                    <th style="width: 8%;">Status</th>
                    <th style="width: 7%;">Confidence</th>
                    <th style="width: 15%;">Case Name</th>
                    <th style="width: 20%;">Context</th>
                </tr>
            </thead>
            <tbody>
"""

    # Data rows
    for term in all_terms:
        case_name = ""
        context = ""

        if term.references and isinstance(term.references, dict):
            case_name = term.references.get("case_name", "")
            context = term.references.get("context", "")

        status_class = f"status-{term.status}"

        html += f"""
                <tr>
                    <td><strong>{term.source_term}</strong></td>
                    <td>{term.target_term or '<em>-</em>'}</td>
                    <td><span class="source-type">{term.source_type}</span></td>
                    <td class="{status_class}">{term.status}</td>
                    <td class="confidence">{term.confidence:.2f}</td>
                    <td>{case_name or '<em>-</em>'}</td>
                    <td class="context">{context or '<em>-</em>'}</td>
                </tr>
"""

    html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    filename = f"glossary_all_{document_id}.html"

    return Response(
        content=html,
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/{document_id}/export/approved/xlsx")
async def export_approved_terms_xlsx(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport tylko zatwierdzonych terminów do XLSX (po zatwierdzeniu).

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Plik XLSX z zatwierdzonymi terminami
    """
    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Pobierz tylko zatwierdzone terminy (approved + edited)
    approved_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id,
        models.Term.status.in_(["approved", "edited"])
    ).order_by(models.Term.source_term).all()

    if not approved_terms:
        raise HTTPException(
            status_code=404,
            detail="No approved terms found for this document"
        )

    # Utwórz workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary - Approved"

    # Style dla nagłówka
    header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style dla danych
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    # Header
    headers = ['Source Term (EN)', 'Target Term (PL)', 'Source Type', 'Status', 'Confidence', 'Original Proposal', 'Case Name', 'Context']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_num, term in enumerate(approved_terms, 2):
        case_name = ""
        context = ""

        if term.references and isinstance(term.references, dict):
            case_name = term.references.get("case_name", "")
            context = term.references.get("context", "")

        # Status color coding
        status_fill = None
        if term.status == "approved":
            status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        elif term.status == "edited":
            status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow

        data = [
            term.source_term,
            term.target_term or "",
            term.source_type,
            term.status,
            f"{term.confidence:.2f}",
            term.original_proposal or "",
            case_name,
            context
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = cell_alignment

            # Apply status color to Status column
            if col_num == 4 and status_fill:
                cell.fill = status_fill

            # Highlight edited terms in Original Proposal column
            if col_num == 6 and value:  # Original Proposal column
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(italic=True, color="856404")

    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # Source Term
    ws.column_dimensions['B'].width = 30  # Target Term
    ws.column_dimensions['C'].width = 15  # Source Type
    ws.column_dimensions['D'].width = 12  # Status
    ws.column_dimensions['E'].width = 12  # Confidence
    ws.column_dimensions['F'].width = 25  # Original Proposal
    ws.column_dimensions['G'].width = 25  # Case Name
    ws.column_dimensions['H'].width = 50  # Context

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"glossary_approved_{document_id}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/{document_id}/export/project-state")
async def export_project_state(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport pełnego stanu projektu do JSON (do późniejszego wznowienia pracy).

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Plik JSON z pełnym stanem projektu
    """
    import json

    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Pobierz wszystkie terminy
    all_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id
    ).order_by(models.Term.source_term).all()

    # Przygotuj dane do eksportu
    terms_data = []
    for term in all_terms:
        terms_data.append({
            "id": term.id,
            "source_term": term.source_term,
            "target_term": term.target_term,
            "original_proposal": term.original_proposal,
            "source_type": term.source_type,
            "confidence": term.confidence,
            "status": term.status,
            "references": term.references,
            "created_at": term.created_at.isoformat() if term.created_at else None,
            "updated_at": term.updated_at.isoformat() if term.updated_at else None,
        })

    project_state = {
        "version": "1.0",
        "document_id": document_id,
        "document_filename": db_document.filename,
        "exported_at": datetime.now().isoformat(),
        "stats": {
            "total": len(all_terms),
            "pending": len([t for t in all_terms if t.status == "pending"]),
            "approved": len([t for t in all_terms if t.status == "approved"]),
            "edited": len([t for t in all_terms if t.status == "edited"]),
            "rejected": len([t for t in all_terms if t.status == "rejected"]),
        },
        "terms": terms_data,
    }

    filename = f"project_state_{document_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    return Response(
        content=json.dumps(project_state, ensure_ascii=False, indent=2),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.post("/{document_id}/import/project-state")
async def import_project_state(document_id: str, db: Session = Depends(get_db)):
    """
    Import stanu projektu z JSON.

    Wymaga przesłania pliku JSON w body requestu.

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Podsumowanie importu
    """
    from fastapi import Request
    # Ten endpoint będzie obsługiwany przez osobną funkcję z File upload
    raise HTTPException(status_code=501, detail="Use /import/project-state-file endpoint with file upload")


@router.put("/{document_id}/restore-terms")
async def restore_terms_state(document_id: str, terms_updates: List[Dict], db: Session = Depends(get_db)):
    """
    Przywróć stan terminów z wcześniej zapisanego projektu.

    Args:
        document_id: ID dokumentu
        terms_updates: Lista terminów z ich stanami
        db: Sesja bazy danych

    Returns:
        Podsumowanie przywrócenia
    """
    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    updated_count = 0
    not_found_count = 0

    for term_data in terms_updates:
        term_id = term_data.get("id")
        if not term_id:
            continue

        # Znajdź termin w bazie
        term = db.query(models.Term).filter(
            models.Term.id == term_id,
            models.Term.document_id == document_id
        ).first()

        if term:
            # Przywróć stan
            if "target_term" in term_data:
                term.target_term = term_data["target_term"]
            if "status" in term_data:
                term.status = term_data["status"]
            if "original_proposal" in term_data:
                term.original_proposal = term_data["original_proposal"]
            term.updated_at = datetime.now()
            updated_count += 1
        else:
            not_found_count += 1

    db.commit()

    return {
        "message": f"Przywrócono stan {updated_count} terminów",
        "updated": updated_count,
        "not_found": not_found_count,
    }


@router.get("/{document_id}/export/approved/html")
async def export_approved_terms_html(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport tylko zatwierdzonych terminów do HTML (po zatwierdzeniu).

    Args:
        document_id: ID dokumentu
        db: Sesja bazy danych

    Returns:
        Plik HTML z zatwierdzonymi terminami
    """
    # Sprawdź czy dokument istnieje
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Pobierz tylko zatwierdzone terminy (approved + edited)
    approved_terms = db.query(models.Term).filter(
        models.Term.document_id == document_id,
        models.Term.status.in_(["approved", "edited"])
    ).order_by(models.Term.source_term).all()

    if not approved_terms:
        raise HTTPException(
            status_code=404,
            detail="No approved terms found for this document"
        )

    # Statystyki
    stats = {
        'total': len(approved_terms),
        'approved': len([t for t in approved_terms if t.status == 'approved']),
        'edited': len([t for t in approved_terms if t.status == 'edited']),
    }

    # Generate HTML
    html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Glossary - Approved Terms - {db_document.filename}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #27ae60;
            border-bottom: 3px solid #27ae60;
            padding-bottom: 10px;
        }}
        .meta {{
            color: #7f8c8d;
            margin-bottom: 20px;
            font-size: 14px;
        }}
        .stats {{
            display: flex;
            gap: 15px;
            margin-bottom: 20px;
            flex-wrap: wrap;
        }}
        .stat {{
            background: #ecf0f1;
            padding: 10px 20px;
            border-radius: 5px;
            font-size: 14px;
        }}
        .stat strong {{
            color: #2c3e50;
            font-size: 18px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #27ae60;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 10px;
            border: 1px solid #ddd;
            vertical-align: top;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #e8f5e9;
        }}
        .status-approved {{
            background-color: #d4edda !important;
            color: #155724;
            font-weight: 600;
        }}
        .status-edited {{
            background-color: #fff3cd !important;
            color: #856404;
            font-weight: 600;
        }}
        .source-type {{
            font-size: 11px;
            padding: 3px 8px;
            border-radius: 3px;
            display: inline-block;
            background: #e9ecef;
        }}
        .confidence {{
            font-weight: 600;
            color: #27ae60;
        }}
        .context {{
            font-size: 12px;
            color: #555;
            max-width: 400px;
        }}
        .original-proposal {{
            font-style: italic;
            color: #856404;
            background: #fff3cd;
            padding: 5px;
            border-radius: 3px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>✅ Glossary - Approved Terms</h1>
        <div class="meta">
            <strong>Document:</strong> {db_document.filename}<br>
            <strong>Document ID:</strong> {document_id}<br>
            <strong>Exported:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>

        <div class="stats">
            <div class="stat">📊 <strong>{stats['total']}</strong> Approved Terms</div>
            <div class="stat" style="background: #d4edda;">✓ <strong>{stats['approved']}</strong> Approved (unchanged)</div>
            <div class="stat" style="background: #fff3cd;">✏️ <strong>{stats['edited']}</strong> Edited</div>
        </div>

        <table>
            <thead>
                <tr>
                    <th style="width: 18%;">Source Term (EN)</th>
                    <th style="width: 18%;">Target Term (PL)</th>
                    <th style="width: 10%;">Source</th>
                    <th style="width: 8%;">Status</th>
                    <th style="width: 7%;">Confidence</th>
                    <th style="width: 14%;">Original Proposal</th>
                    <th style="width: 12%;">Case Name</th>
                    <th style="width: 13%;">Context</th>
                </tr>
            </thead>
            <tbody>
"""

    # Data rows
    for term in approved_terms:
        case_name = ""
        context = ""

        if term.references and isinstance(term.references, dict):
            case_name = term.references.get("case_name", "")
            context = term.references.get("context", "")

        status_class = f"status-{term.status}"
        original_proposal_html = ""
        if term.original_proposal:
            original_proposal_html = f'<span class="original-proposal">{term.original_proposal}</span>'

        html += f"""
                <tr>
                    <td><strong>{term.source_term}</strong></td>
                    <td><strong>{term.target_term or '<em>-</em>'}</strong></td>
                    <td><span class="source-type">{term.source_type}</span></td>
                    <td class="{status_class}">{term.status}</td>
                    <td class="confidence">{term.confidence:.2f}</td>
                    <td>{original_proposal_html or '<em>-</em>'}</td>
                    <td>{case_name or '<em>-</em>'}</td>
                    <td class="context">{context or '<em>-</em>'}</td>
                </tr>
"""

    html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    filename = f"glossary_approved_{document_id}.html"

    return Response(
        content=html,
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# ============================================================================
# GLOSSARY SESSION ENDPOINTS - Save/Resume work progress
# ============================================================================


@router.get("/{document_id}/sessions", response_model=List[GlossarySessionResponse])
async def get_glossary_sessions(document_id: str, db: Session = Depends(get_db)):
    """
    Get all glossary work sessions for a document.

    Args:
        document_id: Document ID
        db: Database session

    Returns:
        List of glossary sessions
    """
    sessions = db.query(models.GlossarySession).filter(
        models.GlossarySession.document_id == document_id
    ).order_by(models.GlossarySession.updated_at.desc()).all()

    return [
        GlossarySessionResponse(
            id=s.id,
            document_id=s.document_id,
            name=s.name,
            current_page=s.current_page,
            status_filter=s.status_filter,
            last_viewed_term_id=s.last_viewed_term_id,
            notes=s.notes,
            is_active=bool(s.is_active),
            created_at=s.created_at,
            updated_at=s.updated_at,
        )
        for s in sessions
    ]


@router.get("/{document_id}/sessions/active", response_model=GlossarySessionResponse)
async def get_active_session(document_id: str, db: Session = Depends(get_db)):
    """
    Get the most recent active session for a document.

    Args:
        document_id: Document ID
        db: Database session

    Returns:
        Active glossary session or 404
    """
    session = db.query(models.GlossarySession).filter(
        models.GlossarySession.document_id == document_id,
        models.GlossarySession.is_active == 1
    ).order_by(models.GlossarySession.updated_at.desc()).first()

    if not session:
        raise HTTPException(status_code=404, detail="No active session found")

    return GlossarySessionResponse(
        id=session.id,
        document_id=session.document_id,
        name=session.name,
        current_page=session.current_page,
        status_filter=session.status_filter,
        last_viewed_term_id=session.last_viewed_term_id,
        notes=session.notes,
        is_active=bool(session.is_active),
        created_at=session.created_at,
        updated_at=session.updated_at,
    )


@router.post("/{document_id}/sessions", response_model=GlossarySessionResponse)
async def create_or_update_session(
    document_id: str,
    session_data: GlossarySessionCreate,
    db: Session = Depends(get_db)
):
    """
    Create a new session or update the active one.

    Args:
        document_id: Document ID
        session_data: Session data
        db: Database session

    Returns:
        Created/updated session
    """
    # Check if document exists
    db_document = db.query(models.Document).filter(
        models.Document.id == document_id
    ).first()
    if not db_document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Find active session or create new one
    session = db.query(models.GlossarySession).filter(
        models.GlossarySession.document_id == document_id,
        models.GlossarySession.is_active == 1
    ).first()

    if session:
        # Update existing session
        session.current_page = session_data.current_page
        session.status_filter = session_data.status_filter
        session.last_viewed_term_id = session_data.last_viewed_term_id
        if session_data.name:
            session.name = session_data.name
        if session_data.notes:
            session.notes = session_data.notes
        session.updated_at = datetime.now()
    else:
        # Create new session
        session = models.GlossarySession(
            id=str(uuid.uuid4()),
            document_id=document_id,
            name=session_data.name or f"Session {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            current_page=session_data.current_page,
            status_filter=session_data.status_filter,
            last_viewed_term_id=session_data.last_viewed_term_id,
            notes=session_data.notes,
            is_active=1,
        )
        db.add(session)

    db.commit()
    db.refresh(session)

    return GlossarySessionResponse(
        id=session.id,
        document_id=session.document_id,
        name=session.name,
        current_page=session.current_page,
        status_filter=session.status_filter,
        last_viewed_term_id=session.last_viewed_term_id,
        notes=session.notes,
        is_active=bool(session.is_active),
        created_at=session.created_at,
        updated_at=session.updated_at,
    )


@router.post("/{document_id}/sessions/{session_id}/complete")
async def complete_session(
    document_id: str,
    session_id: str,
    db: Session = Depends(get_db)
):
    """
    Mark a session as completed (archived).

    Args:
        document_id: Document ID
        session_id: Session ID
        db: Database session

    Returns:
        Success message
    """
    session = db.query(models.GlossarySession).filter(
        models.GlossarySession.id == session_id,
        models.GlossarySession.document_id == document_id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    session.is_active = 0
    session.updated_at = datetime.now()
    db.commit()

    return {"message": "Session marked as completed", "session_id": session_id}


@router.delete("/{document_id}/sessions/{session_id}")
async def delete_session(
    document_id: str,
    session_id: str,
    db: Session = Depends(get_db)
):
    """
    Delete a glossary session.

    Args:
        document_id: Document ID
        session_id: Session ID
        db: Database session

    Returns:
        Success message
    """
    session = db.query(models.GlossarySession).filter(
        models.GlossarySession.id == session_id,
        models.GlossarySession.document_id == document_id
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    db.delete(session)
    db.commit()

    return {"message": "Session deleted", "session_id": session_id}
