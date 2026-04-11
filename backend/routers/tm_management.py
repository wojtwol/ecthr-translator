"""Translation Memory Management API."""

import re
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from pathlib import Path
from lxml import etree
import uuid
import logging

from services.multi_tm_manager import MultiTMManager
from db.database import get_db
from db import models
from config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/tm", tags=["translation-memory"])

# Global TM Manager instance (shared across requests)
_tm_manager: Optional[MultiTMManager] = None


def get_tm_manager() -> MultiTMManager:
    """Get or create TM Manager instance."""
    global _tm_manager
    if _tm_manager is None:
        _tm_manager = MultiTMManager()
        _tm_manager.load_all_from_directory()
    return _tm_manager


class TMInfo(BaseModel):
    """TM information model."""
    name: str
    priority: int
    enabled: bool
    entries_count: int
    file_path: str
    file_type: str = "tmx"


class TMCreateRequest(BaseModel):
    """Request to create/add new TM."""
    name: str
    priority: int = 4
    enabled: bool = True


class TMUpdateRequest(BaseModel):
    """Request to update TM settings."""
    priority: Optional[int] = None
    enabled: Optional[bool] = None


@router.get("/list", response_model=List[TMInfo])
async def list_translation_memories(tm_manager: MultiTMManager = Depends(get_tm_manager)):
    """
    Lista wszystkich pamięci tłumaczeniowych.

    Returns:
        Lista TM z informacjami (name, priority, enabled, entries_count)
    """
    stats = tm_manager.get_stats()

    return [
        TMInfo(
            name=tm_info["name"],
            priority=tm_info["priority"],
            enabled=tm_info["enabled"],
            entries_count=tm_info["entries"],
            file_path=tm_info["file_path"],
            file_type=tm_info.get("file_type", "tmx")
        )
        for tm_info in stats["memories"]
    ]


@router.get("/stats")
async def get_tm_stats(tm_manager: MultiTMManager = Depends(get_tm_manager)):
    """
    Statystyki wszystkich TM.

    Returns:
        Dict z pełnymi statystykami
    """
    return tm_manager.get_stats()


@router.post("/upload")
async def upload_tm_file(
    file: UploadFile = File(...),
    priority: int = 4,
    enabled: bool = True,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Upload i dodanie nowej pamięci tłumaczeniowej (TMX lub TBX).

    Args:
        file: Plik TMX/TBX
        priority: Priorytet 1-5 (1 = najwyższy)
        enabled: Czy włączona

    Returns:
        Info o dodanej TM
    """
    logger.info(f"TM upload request received: filename={file.filename}, content_type={file.content_type}")

    # Validate file extension
    if not file.filename.endswith(('.tmx', '.tbx')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Only .tmx and .tbx files are allowed."
        )

    # Generate unique name and preserve file extension
    original_ext = Path(file.filename).suffix  # .tmx or .tbx
    tm_name = f"uploaded_{uuid.uuid4().hex[:8]}_{Path(file.filename).stem}"

    logger.info(f"Saving TM as: {tm_name}{original_ext}, priority={priority}, enabled={enabled}")

    # Save file with chunk-based reading and size validation
    file_path = settings.tm_path / f"{tm_name}{original_ext}"
    file_size = 0
    chunk_size = 1024 * 1024  # 1MB chunks
    max_size = settings.max_tmx_size_mb * 1024 * 1024  # Convert MB to bytes

    try:
        logger.info(f"Starting chunk-based file save, max size: {settings.max_tmx_size_mb}MB")
        with open(file_path, 'wb') as buffer:
            chunk_count = 0
            while chunk := await file.read(chunk_size):
                chunk_count += 1
                file_size += len(chunk)
                if chunk_count % 10 == 0:  # Log every 10MB
                    logger.info(f"Uploaded {file_size / (1024 * 1024):.2f}MB ({chunk_count} chunks)")

                if file_size > max_size:
                    # Clean up partial file
                    file_path.unlink(missing_ok=True)
                    logger.error(f"File too large: {file_size / (1024 * 1024):.2f}MB > {settings.max_tmx_size_mb}MB")
                    raise HTTPException(
                        status_code=413,
                        detail=f"File too large. Maximum size: {settings.max_tmx_size_mb}MB, uploaded: {file_size / (1024 * 1024):.2f}MB"
                    )
                buffer.write(chunk)

        logger.info(f"File saved successfully: {file_size / (1024 * 1024):.2f}MB ({chunk_count} chunks)")
    except HTTPException:
        raise
    except Exception as e:
        # Clean up on any error
        file_path.unlink(missing_ok=True)
        logger.error(f"Error saving file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")

    # Determine file type from extension
    file_type = "tbx" if original_ext.lower() == ".tbx" else "tmx"

    # Add to TM Manager
    success = tm_manager.add_tm(
        name=tm_name,
        file_path=str(file_path),
        priority=priority,
        enabled=enabled,
        auto_load=True,
        file_type=file_type
    )

    if not success:
        file_path.unlink()  # Clean up
        raise HTTPException(status_code=500, detail="Failed to load TM file")

    # Get info about added TM
    tm = tm_manager.memories[tm_name]

    return {
        "name": tm.name,
        "priority": tm.priority,
        "enabled": tm.enabled,
        "entries_count": len(tm.entries),
        "file_path": tm.file_path,
        "file_type": file_type,
        "file_size_mb": round(file_size / (1024 * 1024), 2),
        "message": f"Successfully uploaded TM with {len(tm.entries)} entries ({file_size / (1024 * 1024):.2f}MB)"
    }


@router.put("/{tm_name}")
async def update_tm_settings(
    tm_name: str,
    update: TMUpdateRequest,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Aktualizuje ustawienia TM (priority, enabled).

    Args:
        tm_name: Nazwa TM
        update: Nowe ustawienia

    Returns:
        Zaktualizowane info o TM
    """
    if tm_name not in tm_manager.memories:
        raise HTTPException(status_code=404, detail=f"TM '{tm_name}' not found")

    # Update priority if provided
    if update.priority is not None:
        if not tm_manager.set_priority(tm_name, update.priority):
            raise HTTPException(status_code=400, detail="Invalid priority value")

    # Update enabled status if provided
    if update.enabled is not None:
        if update.enabled:
            tm_manager.enable_tm(tm_name)
        else:
            tm_manager.disable_tm(tm_name)

    # Return updated info
    tm = tm_manager.memories[tm_name]

    return {
        "name": tm.name,
        "priority": tm.priority,
        "enabled": tm.enabled,
        "entries_count": len(tm.entries),
        "file_path": tm.file_path,
        "message": f"Successfully updated TM '{tm_name}'"
    }


@router.delete("/{tm_name}")
async def delete_tm(
    tm_name: str,
    delete_file: bool = False,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Usuwa pamięć tłumaczeniową.

    Args:
        tm_name: Nazwa TM
        delete_file: Czy usunąć też plik TMX (default: False)

    Returns:
        Potwierdzenie usunięcia
    """
    if tm_name not in tm_manager.memories:
        raise HTTPException(status_code=404, detail=f"TM '{tm_name}' not found")

    tm = tm_manager.memories[tm_name]
    file_path = Path(tm.file_path)

    # Remove from manager
    tm_manager.remove_tm(tm_name)

    # Optionally delete file
    if delete_file and file_path.exists():
        file_path.unlink()
        message = f"Deleted TM '{tm_name}' and removed file"
    else:
        message = f"Deleted TM '{tm_name}' (file kept)"

    return {
        "message": message,
        "deleted_tm": tm_name
    }


@router.post("/{tm_name}/enable")
async def enable_tm(
    tm_name: str,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """Włącza pamięć tłumaczeniową."""
    if not tm_manager.enable_tm(tm_name):
        raise HTTPException(status_code=404, detail=f"TM '{tm_name}' not found")

    return {"message": f"Enabled TM '{tm_name}'"}


@router.post("/{tm_name}/disable")
async def disable_tm(
    tm_name: str,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """Wyłącza pamięć tłumaczeniową."""
    if not tm_manager.disable_tm(tm_name):
        raise HTTPException(status_code=404, detail=f"TM '{tm_name}' not found")

    return {"message": f"Disabled TM '{tm_name}'"}


@router.get("/{tm_name}/download")
async def download_tm(
    tm_name: str,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Pobiera plik TMX.

    Args:
        tm_name: Nazwa TM

    Returns:
        Plik TMX do pobrania
    """
    if tm_name not in tm_manager.memories:
        raise HTTPException(status_code=404, detail=f"TM '{tm_name}' not found")

    tm = tm_manager.memories[tm_name]
    file_path = Path(tm.file_path)

    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"TM file not found: {file_path}")

    return FileResponse(
        path=str(file_path),
        media_type="application/xml",
        filename=f"{tm_name}.tmx"
    )


@router.post("/{tm_name}/save")
async def save_tm_to_file(
    tm_name: str,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Zapisuje TM do pliku TMX (dla runtime updates).

    Args:
        tm_name: Nazwa TM

    Returns:
        Potwierdzenie zapisu
    """
    if not tm_manager.save_tm(tm_name):
        raise HTTPException(status_code=500, detail=f"Failed to save TM '{tm_name}'")

    tm = tm_manager.memories[tm_name]

    return {
        "message": f"Saved TM '{tm_name}' with {len(tm.entries)} entries",
        "file_path": tm.file_path
    }


def _strip_leading_numbering(text: str) -> str:
    """Usuwa numerację z początku segmentu (jak Trados)."""
    if not text:
        return text
    patterns = [
        r'^\d+\s*\.\s*',
        r'^\(\s*[a-zA-Z0-9]+\s*\)\s*',
        r'^§\s*\d+\s*\.?\s*',
        r'^[a-zA-Z0-9]+\s*\)\s*',
    ]
    result = text
    for pattern in patterns:
        result = re.sub(pattern, '', result)
        if result != text:
            break
    return result.strip()


@router.get("/export/{document_id}")
async def export_project_tm(document_id: str, db: Session = Depends(get_db)):
    """
    Eksport pamięci tłumaczeniowej z danego projektu jako TMX.

    Args:
        document_id: ID dokumentu/projektu

    Returns:
        Plik TMX z segmentami z tego projektu
    """
    try:
        logger.info(f"Exporting TM for document: {document_id}")

        db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
        if not db_document:
            raise HTTPException(status_code=404, detail="Document not found")

        segments = (
            db.query(models.Segment)
            .filter(models.Segment.document_id == document_id)
            .filter(models.Segment.target_text.isnot(None))
            .filter(models.Segment.target_text != "")
            .order_by(models.Segment.index)
            .all()
        )

        if not segments:
            raise HTTPException(status_code=404, detail="No translated segments found")

        # Build TMX XML
        tmx = etree.Element("tmx", version="1.4")
        etree.SubElement(tmx, "header",
            creationtool="ECTHR-Translator",
            creationtoolversion="2.0",
            segtype="sentence",
            adminlang="en",
            srclang="en",
            datatype="plaintext",
        )
        body = etree.SubElement(tmx, "body")

        entry_count = 0
        for seg in segments:
            if seg.source_text and seg.target_text:
                src_clean = _strip_leading_numbering(seg.source_text)
                tgt_clean = _strip_leading_numbering(seg.target_text)

                if src_clean and tgt_clean:
                    tu = etree.SubElement(body, "tu")
                    src_tuv = etree.SubElement(tu, "tuv")
                    src_tuv.set("{http://www.w3.org/XML/1998/namespace}lang", "en")
                    src_seg = etree.SubElement(src_tuv, "seg")
                    src_seg.text = src_clean

                    tgt_tuv = etree.SubElement(tu, "tuv")
                    tgt_tuv.set("{http://www.w3.org/XML/1998/namespace}lang", "pl")
                    tgt_seg = etree.SubElement(tgt_tuv, "seg")
                    tgt_seg.text = tgt_clean

                    entry_count += 1

        # Save to file
        export_filename = f"tm_export_{document_id}.tmx"
        file_path = settings.tm_path / export_filename

        tree = etree.ElementTree(tmx)
        tree.write(str(file_path), pretty_print=True, xml_declaration=True, encoding="UTF-8")

        logger.info(f"Exported {entry_count} TM entries for document {document_id}")

        return FileResponse(
            path=str(file_path),
            filename=export_filename,
            media_type="application/xml"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting project TM: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-glossary-xlsx")
async def import_glossary_xlsx(
    file: UploadFile = File(...),
    priority: int = 1,
    tm_manager: MultiTMManager = Depends(get_tm_manager)
):
    """
    Import glosariusza z pliku XLSX.

    Oczekiwany format kolumn:
    - Kolumna 1: Source Term (EN)
    - Kolumna 2: Target Term (PL)

    Args:
        file: Plik XLSX z glosariuszem
        priority: Priorytet 1-5 (domyślnie 1 = najwyższy)

    Returns:
        Info o zaimportowanym glosariuszu
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        from openpyxl import load_workbook
        import tempfile

        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        # Read XLSX
        wb = load_workbook(tmp_path, read_only=True)
        ws = wb.active

        # Parse entries - expect Source Term in col 1, Target Term in col 2
        entries = []
        skipped = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                skipped += 1
                continue

            source_term = str(row[0]).strip() if row[0] else ""
            target_term = str(row[1]).strip() if row[1] else ""

            if source_term and target_term and len(source_term) >= 2:
                entries.append({"source": source_term, "target": target_term})
            else:
                skipped += 1

        wb.close()

        # Clean up temp file
        Path(tmp_path).unlink(missing_ok=True)

        if not entries:
            raise HTTPException(status_code=400, detail="No valid entries found in XLSX file")

        # Convert to TBX and save
        tm_name = f"glossary_{uuid.uuid4().hex[:8]}_{Path(file.filename).stem}"
        tbx_filename = f"{tm_name}.tbx"
        tbx_path = settings.tm_path / tbx_filename

        # Build TBX XML
        martif = etree.Element("martif", type="TBX-Basic")
        martif.set("{http://www.w3.org/XML/1998/namespace}lang", "en")

        header = etree.SubElement(martif, "martifHeader")
        file_desc = etree.SubElement(header, "fileDesc")
        title_stmt = etree.SubElement(file_desc, "titleStmt")
        title = etree.SubElement(title_stmt, "title")
        title.text = f"Imported from {file.filename}"

        text = etree.SubElement(martif, "text")
        body = etree.SubElement(text, "body")

        for entry_data in entries:
            term_entry = etree.SubElement(body, "termEntry")

            lang_en = etree.SubElement(term_entry, "langSet")
            lang_en.set("{http://www.w3.org/XML/1998/namespace}lang", "en")
            tig_en = etree.SubElement(lang_en, "tig")
            term_en = etree.SubElement(tig_en, "term")
            term_en.text = entry_data["source"]

            lang_pl = etree.SubElement(term_entry, "langSet")
            lang_pl.set("{http://www.w3.org/XML/1998/namespace}lang", "pl")
            tig_pl = etree.SubElement(lang_pl, "tig")
            term_pl = etree.SubElement(tig_pl, "term")
            term_pl.text = entry_data["target"]

        tree = etree.ElementTree(martif)
        tree.write(str(tbx_path), pretty_print=True, xml_declaration=True, encoding="UTF-8")

        # Load into TM Manager
        success = tm_manager.add_tm(
            name=tm_name,
            file_path=str(tbx_path),
            priority=priority,
            enabled=True,
            auto_load=True,
            file_type="tbx"
        )

        if not success:
            tbx_path.unlink(missing_ok=True)
            raise HTTPException(status_code=500, detail="Failed to load glossary")

        tm = tm_manager.memories[tm_name]

        logger.info(f"Imported XLSX glossary '{file.filename}': {len(tm.entries)} entries, priority={priority}")

        return {
            "name": tm.name,
            "priority": tm.priority,
            "enabled": tm.enabled,
            "entries_count": len(tm.entries),
            "skipped_rows": skipped,
            "file_type": "tbx",
            "message": f"Imported {len(tm.entries)} glossary entries from XLSX (skipped {skipped} invalid rows)"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing XLSX glossary: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
